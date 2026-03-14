import os
import sys
import subprocess
import re
import xml.etree.ElementTree as ET
import jieba
import math
import csv
import xlrd
import openpyxl
from collections import Counter

def extract_student_info_from_filename(filename):
    """从文件名中提取学号和姓名"""
    parts = filename.split('-')
    if len(parts) >= 6:
        # 学号是倒数第二部分
        student_id = parts[-2]
        # 姓名是最后一部分，去掉文件扩展名
        name_part = parts[-1]
        name = name_part.replace('.doc', '').replace('.docx', '')
        return student_id, name
    return "", filename

def is_gibberish(text):
    """检查文本是否是乱码或图片数据"""
    # 检查是否是base64编码的图片数据
    # JPEG图片的base64编码通常以/9j/开头
    if text.startswith('/9j/') or text.startswith('iVBOR'):
        return True
    
    # 检查是否是长串的base64编码（连续的字母数字字符，没有空格）
    if len(text) > 50:
        # 移除所有空格和换行
        clean_text = text.replace(' ', '').replace('\n', '').replace('\r', '')
        # 如果是连续的base64字符（字母、数字、+、/、=）
        if re.match(r'^[A-Za-z0-9+/=]+$', clean_text):
            return True
    
    # 如果文本主要是英文字母和数字，且长度很长，可能是图片数据
    if len(text) > 100:
        # 检查是否包含大量非中文字符
        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
        if chinese_chars / len(text) < 0.1:  # 中文字符比例小于10%
            return True
    
    # 检查是否包含XML标签
    if text.startswith('<') or text.startswith('w:'):
        return True
    
    return False

def read_word_doc(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        try:
            root = ET.fromstring(content)
            text_content = []
            for elem in root.iter():
                # 获取元素的text属性
                if elem.text and elem.text.strip():
                    text_content.append(elem.text.strip())
                # 获取元素的tail属性（标签后的文本）
                if elem.tail and elem.tail.strip():
                    text_content.append(elem.tail.strip())
            full_text = '\n'.join(text_content)
            
            lines = full_text.split('\n')
            student_answer = []
            in_answer = False
            
            for i, line in enumerate(lines):
                if '学生答案：' in line:
                    in_answer = True
                    # 检查是否直接包含答案内容
                    if len(line) > 5:
                        # 提取"学生答案："后面的内容
                        answer_content = line.split('学生答案：', 1)[-1].strip()
                        if answer_content and not is_gibberish(answer_content):
                            student_answer.append(answer_content)
                    continue
                elif in_answer:
                    # 遇到"正确答案："停止
                    if '正确答案：' in line:
                        in_answer = False
                        break
                    # 跳过乱码和图片数据
                    elif is_gibberish(line):
                        continue
                    # 提取实际答案内容
                    elif line.strip():
                        student_answer.append(line.strip())
            
            if student_answer:
                return '\n'.join(student_answer)
            else:
                return ""
        except ET.ParseError:
            lines = content.split('\n')
            clean_lines = []
            for line in lines:
                line = line.strip()
                if line and len(line) > 1:
                    if not line.startswith('<?') and not line.startswith('<') and not line.startswith('w:'):
                        if re.search(r'[\u4e00-\u9fff]', line):
                            clean_lines.append(line)
            return '\n'.join(clean_lines)
    except Exception as e:
        return f"读取文件出错: {e}"

def extract_student_answers(folder_path):
    folder_path = os.path.abspath(folder_path)
    
    if not os.path.exists(folder_path):
        print(f"文件夹不存在: {folder_path}")
        return
    
    doc_files = [f for f in os.listdir(folder_path) 
                  if (f.endswith('.doc') or f.endswith('.docx')) 
                  and not f.startswith('~$')]  # 过滤掉Word临时文件
    
    if not doc_files:
        print("文件夹中没有找到Word文档")
        return
    
    results = []
    
    for doc_file in doc_files:
        file_path = os.path.join(folder_path, doc_file)
        student_id, name = extract_student_info_from_filename(doc_file)
        content = read_word_doc(file_path)
        results.append((student_id, name, content))
    
    return results

def tokenize(text):
    """使用jieba进行中文分词"""
    words = jieba.lcut(text)
    words = [w for w in words if len(w) > 1 and w.strip()]
    return words

def compute_tf(words):
    """计算词频(TF)"""
    word_count = Counter(words)
    total = len(words)
    tf = {word: count / total for word, count in word_count.items()}
    return tf

def compute_idf(documents):
    """计算逆文档频率(IDF)"""
    n_docs = len(documents)
    idf = {}
    all_words = set()
    for doc in documents:
        all_words.update(doc)
    
    for word in all_words:
        count = sum(1 for doc in documents if word in doc)
        idf[word] = math.log(n_docs / (count + 1)) + 1
    return idf

def compute_tfidf_vector(words, idf):
    """计算TF-IDF向量"""
    tf = compute_tf(words)
    tfidf = {}
    for word, tf_value in tf.items():
        tfidf[word] = tf_value * idf.get(word, 0)
    return tfidf

def cosine_similarity(vec1, vec2):
    """计算两个向量的余弦相似度"""
    common_words = set(vec1.keys()) & set(vec2.keys())
    if not common_words:
        return 0.0
    
    dot_product = sum(vec1[word] * vec2[word] for word in common_words)
    norm1 = math.sqrt(sum(v ** 2 for v in vec1.values()))
    norm2 = math.sqrt(sum(v ** 2 for v in vec2.values()))
    
    if norm1 == 0 or norm2 == 0:
        return 0.0
    
    return dot_product / (norm1 * norm2)

def compute_similarity_matrix(results):
    """计算所有学生答案的相似度矩阵"""
    documents = []
    for student_id, name, content in results:
        words = tokenize(content)
        documents.append(words)
    
    idf = compute_idf(documents)
    
    tfidf_vectors = []
    for words in documents:
        tfidf = compute_tfidf_vector(words, idf)
        tfidf_vectors.append(tfidf)
    
    n = len(results)
    similarity_matrix = [[0.0] * n for _ in range(n)]
    
    for i in range(n):
        for j in range(i + 1, n):
            sim = cosine_similarity(tfidf_vectors[i], tfidf_vectors[j])
            similarity_matrix[i][j] = sim
            similarity_matrix[j][i] = sim
    
    for i in range(n):
        similarity_matrix[i][i] = 1.0
    
    return similarity_matrix

def compute_average_similarity(similarity_matrix, student_idx):
    """计算某个学生在全班的平均相似度（不包括自己）"""
    n = len(similarity_matrix)
    total_sim = sum(similarity_matrix[student_idx][j] for j in range(n) if j != student_idx)
    return total_sim / (n - 1)

def find_excel_file(folder_path):
    """查找文件夹中的Excel文件（排除已更新的文件）"""
    excel_files = []
    for file in os.listdir(folder_path):
        if (file.endswith('.xls') or file.endswith('.xlsx')) and 'updated' not in file.lower():
            excel_files.append(os.path.join(folder_path, file))
    return excel_files

def update_excel_scores(excel_file, student_scores):
    """更新Excel文件中的分数"""
    if excel_file.endswith('.xls'):
        workbook = xlrd.open_workbook(excel_file)
        sheet = workbook.sheet_by_index(0)
        
        # 找到学号和分数列（检查前3行）
        id_col = -1
        score_col = -1
        for row in range(min(3, sheet.nrows)):
            if id_col != -1 and score_col != -1:
                break
            for col in range(sheet.ncols):
                header = str(sheet.cell_value(row, col)).strip()
                if '学号' in header or '学生号' in header:
                    id_col = col
                elif '分数' in header or '得分' in header or '成绩' in header:
                    score_col = col
        
        if id_col == -1 or score_col == -1:
            print("未找到学号或分数列")
            return False
        
        # 创建新的工作簿
        output_file = excel_file.replace('.xls', '_updated.xls')
        # 注意：xlrd只能读，不能写，这里需要使用其他库
        print("提示：.xls文件暂不支持直接修改，建议使用.xlsx文件")
        return False
    
    elif excel_file.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # 找到学号和分数列（检查前3行）
        id_col = -1
        score_col = -1
        for row in range(1, min(4, sheet.max_row + 1)):
            if id_col != -1 and score_col != -1:
                break
            for col in range(1, sheet.max_column + 1):
                header = str(sheet.cell(row=row, column=col).value).strip()
                if '学号' in header or '学生号' in header:
                    id_col = col
                elif '分数' in header or '得分' in header or '成绩' in header:
                    score_col = col
        
        if id_col == -1 or score_col == -1:
            print("未找到学号或分数列")
            return False
        
        # 更新分数
        updated_count = 0
        for row in range(2, sheet.max_row + 1):
            student_id = str(sheet.cell(row=row, column=id_col).value).strip()
            if student_id in student_scores:
                sheet.cell(row=row, column=score_col, value=student_scores[student_id])
                updated_count += 1
        
        output_file = excel_file.replace('.xlsx', '_updated.xlsx')
        workbook.save(output_file)
        print(f"已更新 {updated_count} 个学生的分数到: {output_file}")
        return True
    
    return False

def compute_score_by_similarity(avg_sim, min_sim, max_sim):
    """根据正态分布比例划分区间计算分数"""
    if max_sim == min_sim:
        return 85
    
    diff = max_sim - min_sim
    
    normal_distribution_ratios = [0.067, 0.159, 0.309, 0.691, 0.841, 1.0]
    
    normalized = (avg_sim - min_sim) / diff
    
    scores = [100, 95, 90, 85, 80, 75]
    
    for i, threshold in enumerate(normal_distribution_ratios):
        if normalized <= threshold:
            return scores[i]
    
    return 75

def analyze_similarity(results, folder_path):
    """分析相似度并输出结果"""
    if len(results) < 2:
        print("学生数量不足，无法进行相似度分析")
        return
    
    print("正在计算语义向量...")
    similarity_matrix = compute_similarity_matrix(results)
    
    print("正在计算平均相似度...")
    avg_similarities = []
    for i, (student_id, name, content) in enumerate(results):
        avg_sim = compute_average_similarity(similarity_matrix, i)
        avg_similarities.append((student_id, name, avg_sim))
    
    avg_similarities.sort(key=lambda x: x[2])
    
    min_sim = avg_similarities[0][2]
    max_sim = avg_similarities[-1][2]
    
    print("\n" + "=" * 65)
    print("相似度分析结果（从低到高排序）")
    print("=" * 65)
    print(f"{'排名':<6} {'学号':<15} {'姓名':<10} {'平均相似度':<12} {'分数':<6}")
    print("-" * 65)
    
    student_scores = {}
    for rank, (student_id, name, avg_sim) in enumerate(avg_similarities, 1):
        score = compute_score_by_similarity(avg_sim, min_sim, max_sim)
        print(f"{rank:<6} {student_id:<15} {name:<10} {avg_sim:<12.4f} {score:<6}")
        student_scores[student_id] = score
    print("=" * 65)
    
    # 查找并更新Excel文件
    excel_files = find_excel_file(folder_path)
    if excel_files:
        print("\n找到Excel文件:")
        for excel_file in excel_files:
            print(f"- {os.path.basename(excel_file)}")
            update_excel_scores(excel_file, student_scores)
    else:
        print("\n未找到Excel文件")
    
    return avg_similarities

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python evaluate_class.py <文件夹路径>")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    results = extract_student_answers(folder_path)
    
    if results:
        analyze_similarity(results, folder_path)

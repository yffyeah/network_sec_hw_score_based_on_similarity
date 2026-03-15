import os
import sys
import subprocess
import re
import xml.etree.ElementTree as ET
import jieba
import math
import csv
import xlrd
import xlwt
import openpyxl
from collections import Counter

def extract_student_info_from_filename(filename):
    """从文件名中提取学号和姓名"""
    parts = filename.split('-')
    if len(parts) >= 6:
        # 学号是倒数第二部分
        student_id = parts[-2]
        # 姓名是最后一部分，去掉文件扩展名
        name_part = parts[-1]
        name = name_part.replace('.doc', '').replace('.docx', '')
        return student_id, name
    return "", filename

def is_gibberish(text):
    """检查文本是否是乱码或图片数据"""
    # 检查是否是base64编码的图片数据
    # JPEG图片的base64编码通常以/9j/开头
    if text.startswith('/9j/') or text.startswith('iVBOR'):
        return True
    
    # 检查是否是长串的base64编码（连续的字母数字字符，没有空格）
    if len(text) > 50:
        # 移除所有空格和换行
        clean_text = text.replace(' ', '').replace('\n', '').replace('\r', '')
        # 如果是连续的base64字符（字母、数字、+、/、=）
        if re.match(r'^[A-Za-z0-9+/=]+$', clean_text):
            return True
    
    # 如果文本主要是英文字母和数字，且长度很长，可能是图片数据
    if len(text) > 100:
        # 检查是否包含大量非中文字符
        chinese_chars = len(re.findall(r'[\u4e00-\u9fff]', text))
        if chinese_chars / len(text) < 0.1:  # 中文字符比例小于10%
            return True
    
    # 检查是否包含XML标签
    if text.startswith('<') or text.startswith('w:'):
        return True
    
    return False

def extract_images_from_docx(docx_path):
    """从docx文件中提取图片"""
    import zipfile
    import io
    images = []
    
    try:
        with zipfile.ZipFile(docx_path, 'r') as zip_ref:
            # 遍历zip中的所有文件
            for file_info in zip_ref.infolist():
                if file_info.filename.startswith('word/media/'):
                    # 读取图片数据
                    with zip_ref.open(file_info.filename) as f:
                        image_data = f.read()
                        # 提取图片特征
                        image_feature = extract_image_features(image_data)
                        images.append(image_feature)
    except Exception as e:
        # 打印异常信息以便调试
        print(f"提取图片时出错: {e}")
    
    return images

def extract_image_features(image_data):
    """提取图片特征"""
    # 简单的图片特征提取
    # 1. 图片大小
    size = len(image_data)
    # 2. 颜色分布（通过简单的统计）
    # 对于JPEG图片，前几个字节是固定的，我们可以统计后面的字节分布
    if size > 10:
        # 统计前100个字节的分布
        byte_counts = {}  # 暂时使用空字典，实际可以统计字节频率
    else:
        byte_counts = {}
    
    # 返回特征向量
    return {"size": size, "byte_counts": byte_counts}

def compute_image_similarity(img1_features, img2_features):
    """计算图片相似度"""
    # 基于图片大小的相似度
    size1 = img1_features.get("size", 0)
    size2 = img2_features.get("size", 0)
    
    if size1 == 0 or size2 == 0:
        return 0.0
    
    # 大小相似度（1 - 绝对差/较大值）
    size_similarity = 1.0 - abs(size1 - size2) / max(size1, size2)
    
    return size_similarity

def compute_combined_similarity(text_similarity, image_similarity, text_weight=0.8, image_weight=0.2):
    """计算文本和图片的综合相似度"""
    return text_similarity * text_weight + image_similarity * image_weight

def read_word_doc(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        try:
            root = ET.fromstring(content)
            text_content = []
            for elem in root.iter():
                # 获取元素的text属性
                if elem.text and elem.text.strip():
                    text_content.append(elem.text.strip())
                # 获取元素的tail属性（标签后的文本）
                if elem.tail and elem.tail.strip():
                    text_content.append(elem.tail.strip())
            full_text = '\n'.join(text_content)
            
            lines = full_text.split('\n')
            student_answer = []
            in_answer = False
            
            for i, line in enumerate(lines):
                if '学生答案：' in line:
                    in_answer = True
                    # 检查是否直接包含答案内容
                    if len(line) > 5:
                        # 提取"学生答案："后面的内容
                        answer_content = line.split('学生答案：', 1)[-1].strip()
                        if answer_content and not is_gibberish(answer_content):
                            student_answer.append(answer_content)
                    continue
                elif in_answer:
                    # 遇到"正确答案："停止
                    if '正确答案：' in line:
                        in_answer = False
                        break
                    # 跳过乱码和图片数据
                    elif is_gibberish(line):
                        continue
                    # 提取实际答案内容
                    elif line.strip():
                        student_answer.append(line.strip())
            
            if student_answer:
                return '\n'.join(student_answer)
            else:
                return ""
        except ET.ParseError:
            lines = content.split('\n')
            clean_lines = []
            for line in lines:
                line = line.strip()
                if line and len(line) > 1:
                    if not line.startswith('<?') and not line.startswith('<') and not line.startswith('w:'):
                        if re.search(r'[\u4e00-\u9fff]', line):
                            clean_lines.append(line)
            return '\n'.join(clean_lines)
    except Exception as e:
        return f"读取文件出错: {e}"

def extract_student_answers(folder_path):
    folder_path = os.path.abspath(folder_path)
    
    if not os.path.exists(folder_path):
        print(f"文件夹不存在: {folder_path}")
        return
    
    doc_files = [f for f in os.listdir(folder_path) 
                  if (f.endswith('.doc') or f.endswith('.docx')) 
                  and not f.startswith('~$')]  # 过滤掉Word临时文件
    
    if not doc_files:
        print("文件夹中没有找到Word文档")
        return
    
    results = []
    
    for doc_file in doc_files:
        file_path = os.path.join(folder_path, doc_file)
        student_id, name = extract_student_info_from_filename(doc_file)
        content = read_word_doc(file_path)
        # 提取图片
        images = []
        if file_path.endswith('.docx'):
            images = extract_images_from_docx(file_path)
        else:
            # 对于.doc文件，直接检查文件内容是否包含图片
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    file_content = f.read()
                # 检查是否包含Word图片标签或base64编码的图片
                if any(tag in file_content for tag in ['<w:binData>', '<v:imagedata>', '/9j/', 'iVBOR']):
                    images = [{'size': 1, 'byte_counts': {}}]  # 标记有图片
            except Exception:
                pass
        results.append((student_id, name, content, images))
    
    return results

def tokenize(text):
    """使用jieba进行中文分词"""
    words = jieba.lcut(text)
    words = [w for w in words if len(w) > 1 and w.strip()]
    return words

def compute_tf(words):
    """计算词频(TF)"""
    word_count = Counter(words)
    total = len(words)
    tf = {word: count / total for word, count in word_count.items()}
    return tf

def compute_idf(documents):
    """计算逆文档频率(IDF)"""
    n_docs = len(documents)
    idf = {}
    all_words = set()
    for doc in documents:
        all_words.update(doc)
    
    for word in all_words:
        count = sum(1 for doc in documents if word in doc)
        idf[word] = math.log(n_docs / (count + 1)) + 1
    return idf

def compute_tfidf_vector(words, idf):
    """计算TF-IDF向量"""
    tf = compute_tf(words)
    tfidf = {}
    for word, tf_value in tf.items():
        tfidf[word] = tf_value * idf.get(word, 0)
    return tfidf

def cosine_similarity(vec1, vec2):
    """计算两个向量的余弦相似度"""
    common_words = set(vec1.keys()) & set(vec2.keys())
    if not common_words:
        return 0.0
    
    dot_product = sum(vec1[word] * vec2[word] for word in common_words)
    norm1 = math.sqrt(sum(v ** 2 for v in vec1.values()))
    norm2 = math.sqrt(sum(v ** 2 for v in vec2.values()))
    
    if norm1 == 0 or norm2 == 0:
        return 0.0
    
    return dot_product / (norm1 * norm2)

def compute_similarity_matrix(results):
    """计算所有学生答案的相似度矩阵"""
    documents = []
    all_images = []
    
    for student_id, name, content, images in results:
        words = tokenize(content)
        documents.append(words)
        all_images.append(images)
    
    # 计算文本相似度
    idf = compute_idf(documents)
    
    tfidf_vectors = []
    for words in documents:
        tfidf = compute_tfidf_vector(words, idf)
        tfidf_vectors.append(tfidf)
    
    n = len(results)
    similarity_matrix = [[0.0] * n for _ in range(n)]
    
    for i in range(n):
        for j in range(i + 1, n):
            # 计算文本相似度
            text_sim = cosine_similarity(tfidf_vectors[i], tfidf_vectors[j])

            # 计算图片相似度
            # 情况1：两个文档都有图片，计算图片相似度
            # 情况2：只有一个文档有图片，图片相似度为0%（表示完全不同）
            # 情况3：两个文档都没有图片，图片相似度为100%（不影响综合相似度）
            if all_images[i] and all_images[j]:
                # 计算所有图片对的相似度，取平均值
                image_sims = []
                for img1 in all_images[i]:
                    for img2 in all_images[j]:
                        img_sim = compute_image_similarity(img1, img2)
                        image_sims.append(img_sim)
                if image_sims:
                    image_sim = sum(image_sims) / len(image_sims)
                else:
                    image_sim = 1.0
            elif all_images[i] or all_images[j]:
                # 只有一个文档有图片，图片相似度为0%
                image_sim = 0.0
            else:
                # 两个文档都没有图片，图片相似度设为100%
                image_sim = 1.0

            # 计算综合相似度
            combined_sim = compute_combined_similarity(text_sim, image_sim)

            similarity_matrix[i][j] = combined_sim
            similarity_matrix[j][i] = combined_sim
    
    for i in range(n):
        similarity_matrix[i][i] = 1.0
    
    return similarity_matrix

def compute_average_similarity(similarity_matrix, student_idx):
    """计算某个学生在全班的平均相似度（不包括自己）"""
    n = len(similarity_matrix)
    total_sim = sum(similarity_matrix[student_idx][j] for j in range(n) if j != student_idx)
    return total_sim / (n - 1)

def find_excel_file(folder_path):
    """查找文件夹中的Excel文件（排除已更新的文件）"""
    excel_files = []
    for file in os.listdir(folder_path):
        if (file.endswith('.xls') or file.endswith('.xlsx')) and 'updated' not in file.lower():
            excel_files.append(os.path.join(folder_path, file))
    return excel_files

def update_excel_scores(excel_file, student_scores):
    """更新Excel文件中的分数"""
    if excel_file.endswith('.xls'):
        workbook = xlrd.open_workbook(excel_file)
        sheet = workbook.sheet_by_index(0)
        
        # 找到学号和分数列（检查前3行）
        id_col = -1
        score_col = -1
        for row in range(min(3, sheet.nrows)):
            if id_col != -1 and score_col != -1:
                break
            for col in range(sheet.ncols):
                header = str(sheet.cell_value(row, col)).strip()
                if '学号' in header or '学生号' in header:
                    id_col = col
                elif '分数' in header or '得分' in header or '成绩' in header:
                    score_col = col
        
        if id_col == -1 or score_col == -1:
            print("未找到学号或分数列")
            return False
        
        # 创建新的工作簿
        new_workbook = xlwt.Workbook(encoding='utf-8')
        new_sheet = new_workbook.add_sheet(sheet.name)
        
        # 查找状态列
        status_col = -1
        for row in range(min(3, sheet.nrows)):
            if status_col != -1:
                break
            for col in range(sheet.ncols):
                header = str(sheet.cell_value(row, col)).strip()
                if '状态' in header or '提交状态' in header:
                    status_col = col
                    break
        
        # 复制原有数据，跳过状态为"未交"的行
        new_row_idx = 0
        updated_count = 0
        for row in range(sheet.nrows):
            # 检查是否为"未交"状态（跳过表头行）
            if row > 0 and status_col != -1:
                status_value = str(sheet.cell_value(row, status_col)).strip()
                if '未交' in status_value:
                    continue  # 跳过"未交"的行
            
            # 复制该行数据
            for col in range(sheet.ncols):
                cell_value = sheet.cell_value(row, col)
                new_sheet.write(new_row_idx, col, cell_value)
            
            # 更新分数（跳过表头行）
            if row > 0:
                student_id = str(sheet.cell_value(row, id_col)).strip()
                if student_id in student_scores:
                    new_sheet.write(new_row_idx, score_col, student_scores[student_id])
                    updated_count += 1
            
            new_row_idx += 1
        
        output_file = excel_file.replace('.xls', '_updated.xls')
        new_workbook.save(output_file)
        print(f"已更新 {updated_count} 个学生的分数到: {output_file}")
        return True
    
    elif excel_file.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        # 找到学号和分数列（检查前3行）
        id_col = -1
        score_col = -1
        for row in range(1, min(4, sheet.max_row + 1)):
            if id_col != -1 and score_col != -1:
                break
            for col in range(1, sheet.max_column + 1):
                header = str(sheet.cell(row=row, column=col).value).strip()
                if '学号' in header or '学生号' in header:
                    id_col = col
                elif '分数' in header or '得分' in header or '成绩' in header:
                    score_col = col
        
        if id_col == -1 or score_col == -1:
            print("未找到学号或分数列")
            return False
        
        # 查找状态列
        status_col = -1
        for row in range(1, min(4, sheet.max_row + 1)):
            if status_col != -1:
                break
            for col in range(1, sheet.max_column + 1):
                header = str(sheet.cell(row=row, column=col).value).strip()
                if '状态' in header or '提交状态' in header:
                    status_col = col
                    break
        
        # 创建新的工作簿（使用xlwt保存为.xls格式）
        new_workbook = xlwt.Workbook(encoding='utf-8')
        new_sheet = new_workbook.add_sheet('Sheet1')
        
        # 复制原有数据，跳过状态为"未交"的行，并更新分数
        new_row_idx = 0
        updated_count = 0
        for row in range(1, sheet.max_row + 1):
            # 检查是否为"未交"状态（跳过表头行）
            if row > 1 and status_col != -1:
                status_value = str(sheet.cell(row=row, column=status_col).value).strip()
                if '未交' in status_value:
                    continue  # 跳过"未交"的行
            
            # 复制该行数据
            for col in range(1, sheet.max_column + 1):
                # 如果是分数列且需要更新，则使用新分数
                if col == score_col and row > 1:
                    student_id = str(sheet.cell(row=row, column=id_col).value).strip()
                    if student_id in student_scores:
                        new_sheet.write(new_row_idx, col - 1, student_scores[student_id])
                        updated_count += 1
                        continue
                
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    new_sheet.write(new_row_idx, col - 1, cell_value)
            
            new_row_idx += 1
        
        output_file = excel_file.replace('.xlsx', '_updated.xls')
        new_workbook.save(output_file)
        print(f"已更新 {updated_count} 个学生的分数到: {output_file}")
        return True
    
    return False

def compute_score_by_similarity(avg_sim, min_sim, max_sim):
    """根据正态分布比例划分区间计算分数"""
    if max_sim == min_sim:
        return 85
    
    diff = max_sim - min_sim
    
    normal_distribution_ratios = [0.067, 0.159, 0.309, 0.691, 0.841, 1.0]
    
    normalized = (avg_sim - min_sim) / diff
    
    scores = [100, 95, 90, 85, 80, 75]
    
    for i, threshold in enumerate(normal_distribution_ratios):
        if normalized <= threshold:
            return scores[i]
    
    return 75

def analyze_similarity(results, folder_path):
    """分析相似度并输出结果"""
    if len(results) < 2:
        print("学生数量不足，无法进行相似度分析")
        return
    
    # 统计字数和图片信息
    word_counts = []
    students_with_images = 0
    for student_id, name, content, images in results:
        word_count = len(content)
        word_counts.append(word_count)
        
        # 检查是否包含图片（三种方式）
        # 1. 从docx文件中提取的图片
        if images:
            students_with_images += 1
        # 2. 检查文档内容中是否包含Word图片标签
        elif any(tag in content for tag in ['<w:binData>', '<v:imagedata>']):
            students_with_images += 1
        # 3. 检查文档内容中是否包含base64编码的图片
        elif any(marker in content for marker in ['/9j/', 'iVBOR']):
            students_with_images += 1
    
    # 计算字数分布区间
    min_words = min(word_counts)
    max_words = max(word_counts)
    avg_words = sum(word_counts) / len(word_counts)
    
    # 将字数分为5个区间
    word_ranges = []
    range_size = (max_words - min_words) / 5 if max_words > min_words else 1
    for i in range(5):
        range_start = min_words + i * range_size
        range_end = min_words + (i + 1) * range_size
        count = sum(1 for wc in word_counts if range_start <= wc < range_end)
        if i == 4:  # 最后一个区间包含最大值
            count = sum(1 for wc in word_counts if range_start <= wc <= range_end)
        word_ranges.append((int(range_start), int(range_end), count))
    
    print("\n" + "=" * 65)
    print("字数统计信息")
    print("=" * 65)
    print(f"总人数: {len(results)}")
    print(f"最少字数: {min_words}")
    print(f"最多字数: {max_words}")
    print(f"平均字数: {avg_words:.1f}")
    print(f"有图片的人数: {students_with_images}")
    print("\n字数分布区间:")
    for i, (start, end, count) in enumerate(word_ranges, 1):
        print(f"  区间 {i}: {start} - {end} 字: {count} 人")
    print("=" * 65)
    
    print("\n正在计算语义向量...")
    similarity_matrix = compute_similarity_matrix(results)
    
    print("正在计算平均相似度...")
    avg_similarities = []
    for i, (student_id, name, content, images) in enumerate(results):
        avg_sim = compute_average_similarity(similarity_matrix, i)
        avg_similarities.append((student_id, name, avg_sim))
    
    avg_similarities.sort(key=lambda x: x[2])
    
    min_sim = avg_similarities[0][2]
    max_sim = avg_similarities[-1][2]
    
    print("\n" + "=" * 65)
    print("相似度分析结果（从低到高排序）")
    print("=" * 65)
    print(f"{'排名':<6} {'学号':<15} {'姓名':<10} {'平均相似度':<12} {'分数':<6}")
    print("-" * 65)
    
    student_scores = {}
    for rank, (student_id, name, avg_sim) in enumerate(avg_similarities, 1):
        score = compute_score_by_similarity(avg_sim, min_sim, max_sim)
        print(f"{rank:<6} {student_id:<15} {name:<10} {avg_sim:<12.4f} {score:<6}")
        student_scores[student_id] = score
    print("=" * 65)
    
    # 查找并更新Excel文件
    excel_files = find_excel_file(folder_path)
    chapter_name = ""
    if excel_files:
        print("\n找到Excel文件:")
        for excel_file in excel_files:
            print(f"- {os.path.basename(excel_file)}")
            # 从Excel文件名中提取章节名称
            excel_basename = os.path.basename(excel_file)
            # 移除扩展名
            excel_name_without_ext = os.path.splitext(excel_basename)[0]
            # 移除班级名称（如"23-3"）
            import re
            chapter_match = re.match(r'(.+?)(?:\d{2}-\d+)?$', excel_name_without_ext)
            if chapter_match:
                chapter_name = chapter_match.group(1).strip()
            update_excel_scores(excel_file, student_scores)
    else:
        print("\n未找到Excel文件")
    
    # 最后打印统计信息
    print("\n" + "=" * 65)
    print("统计信息汇总")
    print("=" * 65)
    print(f"总人数: {len(results)}")
    print(f"最少字数: {min_words}")
    print(f"最多字数: {max_words}")
    print(f"平均字数: {avg_words:.1f}")
    print(f"有图片的人数: {students_with_images}")
    print("\n字数分布区间:")
    for i, (start, end, count) in enumerate(word_ranges, 1):
        print(f"  区间 {i}: {start} - {end} 字: {count} 人")
    print("=" * 65)
    
    # 将汇总信息写入文件
    folder_name = os.path.basename(folder_path)
    if chapter_name:
        summary_file = os.path.join(folder_path, f"summary_{chapter_name}_{folder_name}.txt")
    else:
        summary_file = os.path.join(folder_path, f"summary_{folder_name}.txt")
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("=" * 65 + "\n")
        f.write("统计信息汇总\n")
        f.write("=" * 65 + "\n")
        f.write(f"总人数: {len(results)}\n")
        f.write(f"最少字数: {min_words}\n")
        f.write(f"最多字数: {max_words}\n")
        f.write(f"平均字数: {avg_words:.1f}\n")
        f.write(f"有图片的人数: {students_with_images}\n")
        f.write("\n字数分布区间:\n")
        for i, (start, end, count) in enumerate(word_ranges, 1):
            f.write(f"  区间 {i}: {start} - {end} 字: {count} 人\n")
        f.write("=" * 65 + "\n")
    
    print(f"\n汇总信息已保存到: {summary_file}")
    
    return avg_similarities

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用方法: python evaluate_class.py <文件夹路径>")
        sys.exit(1)
    
    folder_path = sys.argv[1]
    results = extract_student_answers(folder_path)
    
    if results:
        analyze_similarity(results, folder_path)
